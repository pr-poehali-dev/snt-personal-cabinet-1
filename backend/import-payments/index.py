"""
Импорт оплат из Excel для администратора.
POST / — принимает base64-encoded xlsx файл, парсит и сохраняет оплаты в БД.
"""
import json
import os
import base64
import hashlib
from io import BytesIO
from datetime import datetime
import psycopg2
import openpyxl


def get_conn():
    return psycopg2.connect(os.environ["DATABASE_URL"])


def handler(event: dict, context) -> dict:
    headers = {
        "Access-Control-Allow-Origin": "*",
        "Access-Control-Allow-Methods": "POST, OPTIONS",
        "Access-Control-Allow-Headers": "Content-Type",
    }

    if event.get("httpMethod") == "OPTIONS":
        return {"statusCode": 200, "headers": headers, "body": ""}

    body = json.loads(event.get("body") or "{}")
    file_b64 = body.get("file")
    if not file_b64:
        return {"statusCode": 400, "headers": headers, "body": json.dumps({"error": "Файл не передан"})}

    file_bytes = base64.b64decode(file_b64)
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb.active

    headers_row = [str(c.value or "").strip().lower() for c in next(ws.iter_rows(min_row=1, max_row=1))]

    def col(name_variants):
        for v in name_variants:
            if v in headers_row:
                return headers_row.index(v)
        return None

    idx_date = col(["дата", "date"])
    idx_amount = col(["сумма", "amount", "сумма оплаты"])
    idx_plot = col(["участок", "plot", "номер участка"])
    idx_service = col(["услуга", "service", "категория", "category"])

    if None in (idx_date, idx_amount, idx_plot, idx_service):
        missing = []
        if idx_date is None: missing.append("дата")
        if idx_amount is None: missing.append("сумма")
        if idx_plot is None: missing.append("участок")
        if idx_service is None: missing.append("услуга")
        return {"statusCode": 400, "headers": headers, "body": json.dumps({
            "error": f"Не найдены колонки: {', '.join(missing)}"
        })}

    conn = get_conn()
    cur = conn.cursor()

    imported = 0
    skipped = 0
    errors = []

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        try:
            date_val = row[idx_date]
            amount_val = row[idx_amount]
            plot_val = str(row[idx_plot] or "").strip().lower()
            service_val = str(row[idx_service] or "").strip()

            if not date_val or not amount_val or not plot_val or not service_val:
                skipped += 1
                continue

            if isinstance(date_val, datetime):
                payment_date = date_val.date()
            else:
                payment_date = datetime.strptime(str(date_val).strip(), "%d.%m.%Y").date()

            amount = float(str(amount_val).replace(",", ".").replace(" ", ""))

            cur.execute(
                "SELECT id FROM users WHERE login = %s AND role = 'user'",
                (plot_val,)
            )
            if not cur.fetchone():
                errors.append(f"Строка {i}: участок '{plot_val}' не найден")
                skipped += 1
                continue

            cur.execute(
                "INSERT INTO payments (plot, service, amount, payment_date) VALUES (%s, %s, %s, %s)",
                (plot_val, service_val, amount, payment_date)
            )
            imported += 1

        except Exception as e:
            errors.append(f"Строка {i}: {str(e)}")
            skipped += 1

    conn.commit()
    cur.close()
    conn.close()

    return {"statusCode": 200, "headers": headers, "body": json.dumps({
        "imported": imported,
        "skipped": skipped,
        "errors": errors[:10]
    }, ensure_ascii=False)}
